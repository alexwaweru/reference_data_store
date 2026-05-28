"""
Sectors & Industries reference-data fetcher.

Pulls NAICS 2022 (2-6 digit) from the US Census Bureau, parses the XLSX, and
emits a hierarchical JSON with a stable shape — codes, titles, parent links.

Source:
    https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx

Outputs (./outputs/):
    sectors.json      — 20 top-level sectors (2-digit codes; some are ranges e.g. "31-33")
    industries.json   — ~1,000 nodes across all 5 NAICS levels (sector, subsector,
                        industry group, NAICS industry, national industry), each
                        with `parent_code` so consumers can reconstruct the tree.

Why both? sectors.json is the small list that's typically shown in dropdown
filters. industries.json is the full taxonomy for tagging job posts / companies
at any granularity. They use the same shape (`code`, `title`, `level`,
`parent_code`) so the same reader code works for both.

Industry classifications don't require embeddings — codes are exact strings,
and the taxonomy is small enough that exact / prefix matching is fast.

Usage:
    python sectors/fetch.py
    python sectors/fetch.py --force
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger("sectors.fetch")

ROOT = Path(__file__).parent
OUTPUTS = ROOT / "outputs"
CACHE = ROOT / ".cache"

NAICS_URL = "https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx"
LOCAL_XLSX = CACHE / "naics_2022.xlsx"

MAX_RETRIES = 3
RETRY_DELAY_S = 5
FETCH_TIMEOUT_S = 120

LEVEL_BY_LEN = {
    2: "sector",          # also matches range codes like "31-33"
    3: "subsector",
    4: "industry_group",
    5: "naics_industry",
    6: "national_industry",
}


def download_xlsx(force: bool) -> Path:
    CACHE.mkdir(parents=True, exist_ok=True)
    if LOCAL_XLSX.exists() and not force:
        logger.info("Using cached %s (pass --force to re-download)", LOCAL_XLSX)
        return LOCAL_XLSX

    logger.info("Downloading %s", NAICS_URL)
    # Census Bureau rejects the default Python user-agent (HTTP 403); send a
    # browser-ish one. This is a public, anonymous download — no auth involved.
    request = urllib.request.Request(
        NAICS_URL,
        headers={"User-Agent": "Mozilla/5.0 (compatible; reference-data-fetcher)"},
    )
    last_err: Exception | None = None
    for attempt in range(MAX_RETRIES):
        try:
            with urllib.request.urlopen(request, timeout=FETCH_TIMEOUT_S) as r:  # noqa: S310
                LOCAL_XLSX.write_bytes(r.read())
            return LOCAL_XLSX
        except Exception as err:  # noqa: BLE001
            last_err = err
            if attempt < MAX_RETRIES - 1:
                logger.warning(
                    "Attempt %d/%d failed: %s. Retrying in %ds...",
                    attempt + 1,
                    MAX_RETRIES,
                    err,
                    RETRY_DELAY_S,
                )
                time.sleep(RETRY_DELAY_S)
    raise RuntimeError("Failed to download NAICS xlsx") from last_err


def parse_xlsx(path: Path) -> list[dict[str, Any]]:
    """
    Walk the workbook and emit one record per code. The hierarchy is encoded in
    the code length (NAICS rule), with a wrinkle: a few sectors are ranges
    ("31-33", "44-45", "48-49") — track which 2-digit prefixes each range owns
    so we can resolve a 3-digit subsector's parent correctly.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    raw: list[tuple[str, str]] = []
    for row in ws.iter_rows(values_only=True):
        code, title = row[1], row[2]
        if code is None or title is None:
            continue
        code_s = str(code).strip()
        title_s = str(title).strip()
        if not code_s or not title_s or code_s.lower().startswith("2022 naics"):
            continue
        raw.append((code_s, title_s))

    # Build prefix → sector_code map (handles ranges).
    sector_by_prefix: dict[str, str] = {}
    for code_s, _ in raw:
        if "-" in code_s:  # e.g. "31-33"
            lo, hi = code_s.split("-")
            for p in range(int(lo), int(hi) + 1):
                sector_by_prefix[f"{p:02d}"] = code_s
        elif len(code_s) == 2:
            sector_by_prefix[code_s] = code_s

    records: list[dict[str, Any]] = []
    for code_s, title_s in raw:
        if "-" in code_s or len(code_s) == 2:
            level = "sector"
            parent = None
        else:
            level = LEVEL_BY_LEN.get(len(code_s))
            if level is None:
                logger.warning("Skipping code with unexpected length: %s", code_s)
                continue
            if level == "subsector":
                parent = sector_by_prefix.get(code_s[:2])
            else:
                parent = code_s[:-1]

        records.append(
            {
                "code": code_s,
                "title": title_s,
                "level": level,
                "parent_code": parent,
            }
        )

    return records


def write_json(path: Path, payload: Any) -> None:
    logger.info("Writing %s (%d records)", path.name, len(payload))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def should_build(path: Path, force: bool) -> bool:
    if force:
        return True
    if path.exists():
        logger.info("Skipping %s (already exists; pass --force to rebuild)", path.name)
        return False
    return True


def run(force: bool) -> None:
    sectors_path = OUTPUTS / "sectors.json"
    industries_path = OUTPUTS / "industries.json"

    if not should_build(sectors_path, force) and not should_build(industries_path, force):
        return

    xlsx = download_xlsx(force=force)
    records = parse_xlsx(xlsx)

    sectors = [r for r in records if r["level"] == "sector"]
    write_json(sectors_path, sectors)
    write_json(industries_path, records)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch + normalize NAICS sectors & industries.")
    parser.add_argument("--force", action="store_true", help="Re-download and rebuild outputs.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    args = parse_args(argv)
    run(force=args.force)
    return 0


if __name__ == "__main__":
    sys.exit(main())
